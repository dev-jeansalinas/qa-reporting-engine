import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker


def generar_excel_snva(
    modulo,
    cliente,
    tester,
    descripcion,
    version,
    inicio_pruebas,
    fin_pruebas,
    casos_prueba,
    logo_path,
    ruta_final,
    **kwargs,
):
    """
    Genera el archivo Excel de Plan de Pruebas con formato SNVA.
    Recibe logo_path y ruta_final para trabajar con la estructura de carpetas.
    """

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Plan de Pruebas"

    # --- Configuración de Estilos ---
    color_turquesa = PatternFill(
        start_color="00DDB3", end_color="00DDB3", fill_type="solid"
    )
    color_purpura = PatternFill(
        start_color="8A5CF5", end_color="8A5CF5", fill_type="solid"
    )
    negrita = Font(bold=True, name="Arial", size=11)
    fuente_arial = Font(name="Arial", size=11)
    fuente_blanca = Font(name="Arial", size=11, color="FFFFFF")
    fuente_blanca_negrita = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bordes = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Encabezado Principal (Logo y Metadatos) ---
    hoja.merge_cells("D4:E8")
    if os.path.exists(logo_path):
        try:
            img = Image(logo_path)
            img.anchor = TwoCellAnchor(
                editAs="twoCell",
                _from=AnchorMarker(col=3, colOff=0, row=3, rowOff=0),
                to=AnchorMarker(col=5, colOff=0, row=8, rowOff=0),
            )
            hoja.add_image(img)
        except Exception as e:
            print(f"[-] Error al procesar el logo: {e}")

    # Estructura de celdas del encabezado
    hoja.merge_cells("G4:I4")  # Cliente
    hoja.merge_cells("F5:F7")  # Etiqueta Descripción
    hoja.merge_cells("G5:I7")  # Valor Descripción
    hoja.merge_cells("G8:K8")  # Tester

    # Títulos y Valores
    hoja["F4"], hoja["G4"] = "Cliente", cliente
    hoja["F5"], hoja["G5"] = "Descripción", descripcion
    hoja["F8"], hoja["G8"] = "Tester", tester

    # Sección Versión y Fechas
    metadatos = [
        ("J4", "Módulo", modulo),
        ("J5", "Versión", version),
        ("J6", "Fecha Inicio", inicio_pruebas),
        ("J7", "Fecha Fin", fin_pruebas),
    ]

    for r, titulo, valor in metadatos:
        hoja[r] = titulo
        hoja["K" + r[1:]] = int(valor) if str(valor).isdigit() else valor
        hoja[r].font = negrita

    # Aplicar estilos al encabezado (Filas 4 a 8)
    for fila in hoja.iter_rows(min_row=4, max_row=8, min_col=4, max_col=11):
        for celda in fila:
            celda.border = bordes
            celda.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True, indent=1
            )
            if celda.column == 6 or (celda.column == 10 and celda.row in [4, 5, 6, 7]):
                celda.fill, celda.font = color_purpura, fuente_blanca_negrita
            else:
                celda.font = fuente_arial

    # --- Título de Sección ---
    hoja.merge_cells("B10:N11")
    hoja["B10"].value = "Plan de pruebas"
    hoja["B10"].font = Font(bold=True, size=12, name="Arial")
    hoja["B10"].alignment = centrado

    # --- Cabecera de la Tabla ---
    titulos = [
        "ID",
        "TIPO DE PRUEBA",
        "CASO DE PRUEBA",
        "DESCRIPCIÓN",
        "PASOS",
        "DATOS DE INGRESO",
        "RESULTADO ESPERADO",
        "RESULTADO OBTENIDO",
        "BUG",
        "FECHA DE INICIO",
        "ESTADO",
    ]

    col_ptr = 2
    for texto in titulos:
        # Lógica de celdas combinadas para Descripción y Estado
        ancho_header = 2 if texto in ["DESCRIPCIÓN", "ESTADO"] else 1
        if ancho_header > 1:
            hoja.merge_cells(
                start_row=13, start_column=col_ptr, end_row=13, end_column=col_ptr + 1
            )

        celda = hoja.cell(row=13, column=col_ptr, value=texto)
        celda.fill, celda.font, celda.border, celda.alignment = (
            color_turquesa,
            negrita,
            bordes,
            centrado,
        )

        # Pintar la celda gemela de la combinación para los bordes
        if ancho_header > 1:
            gemela = hoja.cell(row=13, column=col_ptr + 1)
            gemela.fill, gemela.border = color_turquesa, bordes

        col_ptr += ancho_header

    # --- Cuerpo de la Tabla ---
    dv = DataValidation(
        type="list",
        formula1='"N.A,Por hacer,En proceso,Bloqueado,Exitoso,Fallido,Solucionado"',
        allow_blank=True,
    )
    hoja.add_data_validation(dv)

    fila_actual = 14
    for i, caso in enumerate(casos_prueba):
        es_par = i % 2 == 0
        color_fondo = color_purpura if es_par else None
        fuente_actual = fuente_blanca if es_par else fuente_arial

        hoja.row_dimensions[fila_actual].height = 30
        hoja.row_dimensions[fila_actual + 1].height = 30

        # Datos a insertar
        datos_fila = [
            f"TC-{str(i+1).zfill(2)}",
            "Funcional",
            caso.get("nombre_caso", ""),
            caso.get("descripcion", ""),
            caso.get("pasos", ""),
            caso.get("datos", "N/A"),
            caso.get("resultado_esperado", ""),
            caso.get("resultado_obtenido", ""),
            caso.get("bug", ""),
            datetime.datetime.now().strftime("%d/%m/%Y"),
        ]

        # Insertar datos con merge
        col_idx = 2
        for idx, val in enumerate(datos_fila):
            es_desc = (
                idx == 3
            )  # Columna descripción es la única que se une doble en datos
            ancho = 2 if es_desc else 1
            hoja.merge_cells(
                start_row=fila_actual,
                start_column=col_idx,
                end_row=fila_actual + 1,
                end_column=col_idx + ancho - 1,
            )

            # Estilo para todas las celdas del bloque merge
            for r_off in [0, 1]:
                for c_off in range(ancho):
                    c = hoja.cell(row=fila_actual + r_off, column=col_idx + c_off)
                    if r_off == 0 and c_off == 0:
                        c.value = val
                    c.alignment, c.border, c.font = centrado, bordes, fuente_actual
                    if color_fondo:
                        c.fill = color_fondo
            col_idx += ancho

        # Columnas de Estado (M y N)
        for col in [13, 14]:
            val_est = "Por hacer" if col == 13 else ""
            val_fec = datetime.datetime.now().strftime("%d/%m/%Y") if col == 13 else ""
            c_sel = hoja.cell(row=fila_actual, column=col, value=val_est)
            c_fec = hoja.cell(row=fila_actual + 1, column=col, value=val_fec)
            dv.add(c_sel)
            for r in [c_sel, c_fec]:
                r.alignment, r.border, r.font = centrado, bordes, fuente_actual
                if color_fondo:
                    r.fill = color_fondo

        fila_actual += 2

    # --- Formato Condicional de Estados ---
    reglas = [
        ('"Exitoso"', "C6EFCE", "006100"),
        ('"Solucionado"', "C6EFCE", "006100"),
        ('"Fallido"', "FFC7CE", "9C0006"),
        ('"Bloqueado"', "FFC7CE", "9C0006"),
        ('"Por hacer"', "D3D3D3", "000000"),
        ('"En proceso"', "FFEB9C", "9C6500"),
        ('"N.A"', "D3D3D3", "000000"),
    ]

    rango_estado = f"M14:N{fila_actual}"
    for formula, bg, fg in reglas:
        hoja.conditional_formatting.add(
            rango_estado,
            CellIsRule(
                operator="equal",
                formula=[formula],
                stopIfTrue=True,
                fill=PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
                font=Font(color=fg),
            ),
        )

    # --- Ajustes Finales de Columnas ---
    anchos = {
        "B": 8,
        "C": 15,
        "D": 25,
        "E": 25,
        "F": 25,
        "G": 40,
        "H": 25,
        "I": 30,
        "J": 30,
        "K": 15,
        "L": 18,
        "M": 18,
        "N": 18,
    }
    for col, width in anchos.items():
        hoja.column_dimensions[col].width = width

    # Guardar en la ruta final definida por run.py
    libro.save(ruta_final)
